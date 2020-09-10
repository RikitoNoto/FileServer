import openpyxl
import os
import CensusuTract

path = os.path.join("D://Download", "python", "automatestuff-ja-master", "automatestuff-ja-master", "ch12", "censuspopdata.xlsx")
book = openpyxl.load_workbook(path)

sheet = book[book.sheetnames[0]]
census = []
pop_count = {}

for i in range(2, sheet.max_row):
    census.append(CensusuTract.CensusTract(sheet.cell(i, 1).value, sheet.cell(i, 2).value, sheet.cell(i, 3).value, sheet.cell(i, 4).value))


for cen in census:
    try:
        pop_count[cen.county] += cen.pop
    except KeyError:
        pop_count[cen.county] = cen.pop

for string in pop_count:
    print(string+" : "+str(pop_count[string]))